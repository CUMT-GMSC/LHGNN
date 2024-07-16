import config
from utils import print_summary
from utils import DataLoader
from utils import HNHN_params
import torch
import dhg
from Aggregator import Aggregator
from HEBatch import HEBatchGenerator
from tqdm import tqdm
from train_test import train_batch
from train_test import test
import pandas as pd
from openpyxl import load_workbook
from LHGNN import LHGNN

device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

# 超参数设置
config = config.parse()
print_summary(config)

#数据集中的节点个数
node_num_dict = {
    'email-Enron': 143,
    'email-Eu': 979,
    'contact-high-school': 327,
    'contact-primary-school': 242,
    'NDC-classes': 1149,
    'congress-bills': 1718
}
num_v = node_num_dict[config.dataset_name]

# 数据集：训练集、验证集、测试集， 60，20，20
train_pos_edges, train_edges, train_labels, valid_edges, valid_labels, test_edges, test_labels = DataLoader(config.dataset_name, config.neg_sample)

# 用于生成节点嵌入的超图结构, 以及HNHN的参数
Hg = dhg.Hypergraph(num_v=num_v, e_list=train_pos_edges, merge_op='sum', device=device)

if config.hgnn_type == 'LHGNN':
    De_alpha, De_sum, Dv_beta, Dv_sum = HNHN_params(Hg, config.edge_alpha, config.node_beta)

if __name__ == "__main__":
    # 新建excel
    # 列标题
    column_titles = ['dataset_name', 'time', 'auc', 'ap', 'acc', 'f1', 'TP', 'TN', 'FP', 'FN', 'loss']
    # 创建一个空的数据帧，只包含列标题
    df = pd.DataFrame(columns=column_titles)
    with pd.ExcelWriter('results.xlsx', engine='xlsxwriter') as writerexcel:
        df.to_excel(writerexcel, index=False)

    for time in range(1, 11):
        # 均匀分布
        X = torch.rand(num_v, config.initial_dim).to(device)
        # 正态分布
        # X = torch.randn(num_v, config.initial_dim).to(device)
        print(f'============================================ 第{time}次训练和测试 ==================================================')
        if config.hgnn_type == 'LHGNN':
            rep_model = LHGNN(
                in_channels=config.initial_dim,
                hid_channels=config.hidden_dim,
                output_channels=config.output_dim,
                De=De_alpha,
                De_sum=De_sum,
                Dv=Dv_beta,
                Dv_sum=Dv_sum,
                drop_rate=config.drop_out
            )

        rep_model.to(device)
        if config.agg_mode == 'mean_std' or config.agg_mode == 'norm_maxmin':
            output_dim = config.output_dim * 2
        else:
            output_dim = config.output_dim
        agg_model = Aggregator(layers=[output_dim, int(output_dim / 2), 2], attention_head=config.n_head, output_dim=config.output_dim)
        agg_model.to(device)
        train_batch_loader = HEBatchGenerator(train_edges, train_labels, config.batch_size, device)

        print(f'============================================ Start Train ==================================================')
        best_valid_accuracy = 0.0
        best_epoch = 0
        patience_epoch = 0
        best_epoch_loss = 0

        for epoch in tqdm(range(1, config.epoch + 1), leave=False):
            loss_list = []
            while True:
                edges, labels, is_last = train_batch_loader.next()
                loss = train_batch(edges, labels, rep_model, agg_model, config.lr, Hg, X, config.agg_mode, config.weight_decay)
                loss_list.append(loss)
                if is_last:
                    break
            epoch_loss = sum(loss_list) / len(loss_list)
            # 每一个epoch结束对验证集做测试，看看准确率多少
            auc, ap, acc, f1, TP, TN, FP, FN = test(valid_edges, valid_labels, Hg, rep_model, agg_model, X, config.agg_mode)
            if acc > best_valid_accuracy:
                print(f'epoch={epoch}, auc={auc}, ap={ap}, acc={acc}, f1={f1}, TP={TP}, TN={TN}, FP={FP}, FN={FN}, loss={epoch_loss}')
                best_valid_accuracy = acc
                best_epoch = epoch
                patience_epoch = 0
                best_epoch_loss = epoch_loss
                # 保存模型
                torch.save(rep_model.state_dict(), f'./checkpoints/rep_model.pkt')
                torch.save(agg_model.state_dict(), f'./checkpoints/agg_model.pkt')
            else:
                patience_epoch += 1
                if patience_epoch >= config.patient_epoch:
                    print('=== Early Stopping===')
                    break
        print(f'best_valid_accuracy: {best_valid_accuracy}, best_epoch: {best_epoch}')
        print(f'============================================ Start Test ==================================================')
        if config.hgnn_type == 'LHGNN':
            rep_model = LHGNN(
                in_channels=config.initial_dim,
                hid_channels=config.hidden_dim,
                output_channels=config.output_dim,
                De=De_alpha,
                De_sum=De_sum,
                Dv=Dv_beta,
                Dv_sum=Dv_sum,
                drop_rate=config.drop_out
            )

        rep_model.to(device)
        if config.agg_mode == 'mean_std' or config.agg_mode == 'norm_maxmin':
            output_dim = config.output_dim * 2
        else:
            output_dim = config.output_dim
        agg_model = Aggregator(layers=[output_dim, int(output_dim / 2), 2], attention_head=config.n_head, output_dim=config.output_dim)
        agg_model.to(device)
        rep_net_path = './checkpoints/rep_model.pkt'
        agg_net_path = './checkpoints/agg_model.pkt'
        rep_model.load_state_dict(torch.load(rep_net_path))
        agg_model.load_state_dict(torch.load(agg_net_path))
        auc, ap, acc, f1, TP, TN, FP, FN = test(test_edges, test_labels, Hg, rep_model, agg_model, X, config.agg_mode)
        print(f'auc={auc}, ap={ap}, acc={acc}, f1={f1}, TP={TP}, TN={TN}, FP={FP}, FN={FN}')
        data = [config.dataset_name, time, auc, ap, acc, f1, TP, TN, FP, FN, best_epoch_loss]
        file_path = 'results.xlsx'
        book = load_workbook(file_path)
        writer1 = pd.ExcelWriter(file_path, engine='openpyxl')
        writer1.book = book
        df = pd.DataFrame([data])
        # 将数据追加到 Excel 文件的 Sheet1（如果不存在该工作表，会自动创建）
        df.to_excel(writer1, index=False, header=False, sheet_name='Sheet1',
                    startrow=writer1.sheets['Sheet1'].max_row)
        # 保存更改
        writer1.save()




